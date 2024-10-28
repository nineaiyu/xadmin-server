import json
import os
from tempfile import NamedTemporaryFile, mktemp

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from rest_framework import serializers
from rest_framework.utils import encoders

from common.core.serializers import BasePrimaryKeyRelatedField
from .base import BaseFileRenderer


class ExcelFileRenderer(BaseFileRenderer):
    media_type = "application/xlsx"
    format = "xlsx"

    wb = None
    ws = None
    row_count = 0

    def initial_writer(self):
        self.wb = Workbook()
        self.ws = self.wb.active

    def write_row(self, row):
        self.row_count += 1
        self.ws.row_dimensions[self.row_count].height = 20
        column_count = 0
        for cell_value in row:
            # 处理非法字符
            column_count += 1
            cell_value = ILLEGAL_CHARACTERS_RE.sub(r'', str(cell_value))
            cell = self.ws.cell(row=self.row_count, column=column_count, value=str(cell_value))
            # 设置单元格格式为纯文本, 防止执行公式
            cell.data_type = 's'

    def format_values(self, data, related=False):
        result = []
        for key, value in data.items():
            if related:
                result.append(f"{value}")
            else:
                result.append(f"{value}({key})")
        return json.loads(json.dumps(result, cls=encoders.JSONEncoder, ensure_ascii=False))

    def add_validation(self, rendered_fields):
        if self.template not in ['import', 'update']:
            return
        validation_data_dict = {}
        w_data = self.wb.create_sheet("data", 1)
        w_data.sheet_state = "hidden"
        # self.ws = self.wb.active
        boolean_choices = {"True": "Yes", "False": "No"}
        for index, field in enumerate(rendered_fields):
            name = field.label
            if field.required:
                name = '*' + name
            name = str(name)
            if isinstance(field, serializers.BooleanField):
                validation_data_dict[name] = self.format_values(boolean_choices)
            if hasattr(field, 'choices'):
                validation_data_dict[name] = self.format_values(getattr(field, 'choices'),
                                                                isinstance(field, BasePrimaryKeyRelatedField))
            if validation_data_dict.get(name) is not None:
                column_letter = get_column_letter(len(validation_data_dict))
                ## 数据验证不支持多选，后期优化
                dv = DataValidation(
                    type="list",
                    formula1=f"{quote_sheetname('data')}!${column_letter}$2:${column_letter}${len(validation_data_dict[name]) + 1}",
                    allow_blank=True,
                )
                self.ws.add_data_validation(dv)
                dv.add(f"{get_column_letter(index + 1)}2:{get_column_letter(index + 1)}1048576")

        w_data.append(list(validation_data_dict.keys()))
        for index, validation_data in enumerate(validation_data_dict.values()):
            for inx, ele in enumerate(validation_data):
                w_data[f"{get_column_letter(index + 1)}{inx + 2}"] = ele

    def after_render(self):
        count = 0
        for col in self.ws.columns:
            max_length = 0
            count += 1
            column = col[0].column_letter
            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            adjusted_width = (max_length + 2) * 1.0
            adjusted_width = 300 if adjusted_width > 300 else adjusted_width
            adjusted_width = 30 if adjusted_width < 30 else adjusted_width
            self.ws.column_dimensions[column].width = adjusted_width

        #         self.wb.save('/tmp/test.xlsx')

        if count:
            row = get_column_letter(count)
            tab = Table(displayName="Table", ref=f"A1:{row}{self.row_count}")
            style = TableStyleInfo(
                name="TableStyleLight13",
                showFirstColumn=True,
                showLastColumn=True,
                showRowStripes=True,
                showColumnStripes=True,
            )
            tab.tableStyleInfo = style
            self.ws.add_table(tab)

    def get_rendered_value(self):
        if os.name == 'nt':
            ## 针对 windows 平台，解决 NamedTemporaryFile 方法 权限异常
            tmp_name = mktemp()
            self.wb.save(tmp_name)
            with open(tmp_name, 'rb') as tmp:
                value = tmp.read()
            os.unlink(tmp_name)
            return value
        else:
            with NamedTemporaryFile() as tmp:
                self.wb.save(tmp.name)
                tmp.seek(0)
                return tmp.read()
